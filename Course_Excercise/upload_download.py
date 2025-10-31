import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

fruit_name = "Apple"
file_path = "C:/Users/hp/Downloads/download.xlsx"
col_name = "price"
new_value = "505"
Dict={}

def file_getupdate(path, search_term, col_name, new_value):
    book = openpyxl.load_workbook(path)
    sheet = book.active
    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=i).value == col_name:
            Dict["col"] = i

    for i in range(1,sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == search_term:
                Dict["row"] = i

    sheet.cell(row=Dict["row"],column=Dict["col"]).value = new_value
    book.save(path)


driver = webdriver.Chrome()
driver.maximize_window()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID, "downloadButton").click()
time.sleep(2)

file_getupdate(file_path, fruit_name, col_name, new_value)

driver.find_element(By.XPATH, "//input[@type='file']").send_keys(file_path)
time.sleep(1)
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[@class='Toastify__toast-body']/div[2]")))
success_message = driver.find_element(By.XPATH, "//div[@class='Toastify__toast-body']/div[2]").text
print(success_message)

price_columnid = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
print(price_columnid)
prices_path = "//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+price_columnid+"-undefined']"
prices_value = driver.find_element(By.XPATH, f"{prices_path}").text
print(prices_value)
time.sleep(2)


