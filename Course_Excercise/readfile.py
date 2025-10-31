import os.path
import openpyxl

header = []

file_path = os.path.abspath("../files/customer_details.xlsx")
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# getting the value of the header in array
for i in range(1, sheet.max_column+1):
    header.append(sheet.cell(row=1, column=i).value)

# inititalizing the dictionary with value as array
data_dict = {}
for head in header:
    data_dict[head] = []

# instead of k it iterates both index and value in enumerate
for i, head in enumerate(header,start=1):
#k=0
#for i in range(1, sheet.max_column+1):
    for j in range(2, sheet.max_row+1):
        data_dict[head].append(sheet.cell(row=j, column=i).value)
    #k = k + 1
print(header)

print(data_dict)
print(data_dict["Name"][0])
print(data_dict["URL"][1])
